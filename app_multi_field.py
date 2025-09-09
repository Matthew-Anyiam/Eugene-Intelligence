from fastapi import FastAPI, File, UploadFile
from fastapi.responses import HTMLResponse, FileResponse
import openpyxl
import os
import uuid
from datetime import datetime

app = FastAPI()

# Enhanced SEC data with more fields
SEC_DATA = {
    "AAPL": {
        "revenue": 94836,
        "net_income": 20721, 
        "total_assets": 352755,
        "cash": 29965,
        "fiscal_year": "2024"
    },
    "COTY": {
        "revenue": 5683,
        "net_income": 126,
        "total_assets": 11234,
        "cash": 287,
        "fiscal_year": "2024"
    }
}

@app.get("/", response_class=HTMLResponse)
def home():
    return """
    <html>
    <head>
        <style>
            body { font-family: Arial; max-width: 900px; margin: 50px auto; padding: 20px; }
            .container { background: #f5f5f5; padding: 30px; border-radius: 10px; }
            h1 { color: #333; }
            .upload-box { background: white; padding: 20px; border-radius: 8px; margin: 20px 0; }
            input, select, button { margin: 10px 0; padding: 10px; font-size: 16px; }
            button { background: #4CAF50; color: white; border: none; cursor: pointer; border-radius: 5px; }
            button:hover { background: #45a049; }
            .results { background: white; padding: 20px; border-radius: 8px; margin-top: 20px; }
            .change { background: #e8f5e9; padding: 10px; margin: 5px 0; border-radius: 4px; }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>🚀 Eugene Intelligence - Multi-Field Updater</h1>
            <div class="upload-box">
                <h3>Step 1: Select Your Excel File</h3>
                <input type="file" id="file" accept=".xlsx">
                
                <h3>Step 2: Select Company</h3>
                <select id="ticker">
                    <option value="AAPL">Apple (AAPL)</option>
                    <option value="COTY">Coty (COTY)</option>
                </select>
                
                <br>
                <button onclick="updateFile()">📊 Update with Latest SEC Data</button>
            </div>
            
            <div id="results" class="results" style="display:none;"></div>
        </div>
        
        <script>
            async function updateFile() {
                const file = document.getElementById('file').files[0];
                if (!file) {
                    alert('Please select an Excel file');
                    return;
                }
                
                const ticker = document.getElementById('ticker').value;
                const formData = new FormData();
                formData.append('file', file);
                
                document.getElementById('results').innerHTML = '<h3>⏳ Processing...</h3>';
                document.getElementById('results').style.display = 'block';
                
                const response = await fetch('/update/' + ticker, {
                    method: 'POST',
                    body: formData
                });
                
                const data = await response.json();
                
                let html = '<h3>✅ Update Complete!</h3>';
                html += '<p><strong>Company:</strong> ' + ticker + '</p>';
                html += '<p><strong>Changes Made:</strong> ' + data.changes_count + '</p>';
                
                if (data.changes.length > 0) {
                    html += '<h4>Updated Fields:</h4>';
                    data.changes.forEach(change => {
                        html += '<div class="change">';
                        html += '<strong>' + change.field + '</strong><br>';
                        html += 'Location: ' + change.location + '<br>';
                        html += 'Old: ' + change.old + ' → New: ' + change.new;
                        html += '</div>';
                    });
                }
                
                if (data.download) {
                    html += '<br><a href="' + data.download + '" download>';
                    html += '<button>📥 Download Updated Excel</button></a>';
                }
                
                document.getElementById('results').innerHTML = html;
            }
        </script>
    </body>
    </html>
    """

@app.post("/update/{ticker}")
async def update_multi_fields(ticker: str, file: UploadFile = File(...)):
    # Save file
    file_id = str(uuid.uuid4())[:8]
    os.makedirs("uploads", exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = f"uploads/{file_id}_{timestamp}_{file.filename}"
    
    with open(path, "wb") as f:
        f.write(await file.read())
    
    # Load workbook
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    
    changes = []
    sec_data = SEC_DATA.get(ticker, {})
    
    # Search patterns for different financial fields
    search_patterns = {
        "revenue": ["revenue", "net revenue", "total revenue", "sales", "net sales"],
        "net_income": ["net income", "net profit", "net earnings", "profit", "earnings"],
        "total_assets": ["total assets", "assets"],
        "cash": ["cash", "cash and cash equivalents"]
    }
    
    # Scan the worksheet for financial fields
    for row in range(1, min(50, ws.max_row + 1)):
        for col in range(1, min(20, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell_text = str(cell.value).lower().strip()
                
                # Check each field pattern
                for field_key, patterns in search_patterns.items():
                    if field_key in sec_data:
                        for pattern in patterns:
                            if pattern in cell_text:
                                # Found a match - update the adjacent cell or cell below
                                # Try cell to the right first
                                target_cell = ws.cell(row=row, column=col+1)
                                
                                # If right cell is empty or not numeric, try cell below
                                if not target_cell.value or not isinstance(target_cell.value, (int, float)):
                                    target_cell = ws.cell(row=row+1, column=col)
                                
                                old_value = target_cell.value
                                new_value = sec_data[field_key]
                                target_cell.value = new_value
                                
                                changes.append({
                                    "field": field_key.replace("_", " ").title(),
                                    "location": f"Row {target_cell.row}, Col {target_cell.column}",
                                    "old": old_value,
                                    "new": new_value
                                })
                                
                                # Only update once per field
                                break
    
    # Save updated file
    updated_path = path.replace('.xlsx', '_UPDATED.xlsx')
    wb.save(updated_path)
    
    return {
        "ticker": ticker,
        "changes_count": len(changes),
        "changes": changes,
        "download": f"/download/{os.path.basename(updated_path)}",
        "sec_data": sec_data
    }

@app.get("/download/{filename}")
async def download(filename: str):
    filepath = f"uploads/{filename}"
    if os.path.exists(filepath):
        return FileResponse(filepath, filename=f"UPDATED_{filename}")
    return {"error": "File not found"}

@app.get("/api/test")
def test():
    return {"status": "API is working", "time": datetime.now().isoformat()}

if __name__ == "__main__":
    import uvicorn
    print("Starting Eugene Intelligence Multi-Field Updater...")
    print("Open http://localhost:8006 in your browser")
    uvicorn.run(app, port=8006, reload=False)
